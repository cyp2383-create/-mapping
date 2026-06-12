"""
数据导出模块 — Excel / CSV / JSON

输出格式：
  Excel: 姓名|职位|公司|城市|技能|联系方式|来源|置信度|HR状态
"""

import json
import csv
from datetime import datetime
from pathlib import Path
from typing import Optional


EXPORT_DIR = Path(__file__).parent.parent / "data" / "exports"

TALENT_COLUMNS = [
    ("name", "姓名"), ("current_company", "当前公司"), ("current_title", "当前职位"),
    ("city", "城市"), ("total_years", "工作年限"), ("skills", "技能标签"),
    ("contact_type", "联系方式类型"), ("contact_value", "联系方式"),
    ("source_platform", "数据来源"), ("source_url", "来源链接"),
    ("confidence", "置信度"), ("confidence_notes", "置信度说明"),
]

JD_COLUMNS = [
    ("title", "职位名称"), ("company", "公司全名"), ("salary", "薪资范围"),
    ("location", "工作地点"), ("experience", "经验要求"), ("education", "学历要求"),
    ("industry", "所属行业"), ("company_size", "公司规模"), ("skills", "技能标签"),
    ("responsibilities", "岗位职责"), ("requirements", "任职要求"), ("bonus", "福利加分"),
    ("source_platform", "数据来源"), ("source_url", "来源链接"),
]


def export_excel_jds(jds: list[dict], filename: Optional[str] = None) -> str:
    """Export JDs to Excel"""
    return _export_excel(jds, JD_COLUMNS, filename, "JD")


def export_excel(talents: list[dict], filename: Optional[str] = None) -> str:
    """Export talents to Excel"""
    return _export_excel(talents, TALENT_COLUMNS, filename, "Talent")


def _export_excel(data: list[dict], columns: list, filename: Optional[str] = None,
                  sheet_name: str = "Data") -> str:
    """导出为 Excel 文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[exporter] openpyxl not installed, falling back to CSV")
        return export_csv(talents, filename)

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    if filename is None:
        filename = f"{sheet_name.lower()}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = EXPORT_DIR / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    high_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    mid_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    low_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")

    for col_idx, (_, header) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = thin_border

    for row_idx, row in enumerate(data, 2):
        for col_idx, (key, _) in enumerate(columns, 1):
            value = row.get(key, "")
            if isinstance(value, (list, dict)): value = json.dumps(value, ensure_ascii=False)
            elif value is None: value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        conf = row.get("confidence", 0.5)
        conf_col = [i for i, (k, _) in enumerate(columns, 1) if k == "confidence"]
        if conf_col:
            cc = ws.cell(row=row_idx, column=conf_col[0])
            if conf >= 0.7: cc.fill = high_fill
            elif conf >= 0.4: cc.fill = mid_fill
            else: cc.fill = low_fill

    for col_idx in range(1, len(columns)+1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(data)+1}"

    wb.save(str(filepath))
    print(f"[exporter] {sheet_name}: {len(data)} rows -> {filepath}")
    return str(filepath)


def export_csv(data: list[dict], columns: list = None, filename: Optional[str] = None) -> str:
    if columns is None: columns = TALENT_COLUMNS
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    if filename is None:
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    filepath = EXPORT_DIR / filename
    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([h for _, h in columns])
        for row in data:
            vals = []
            for key, _ in columns:
                v = row.get(key, "")
                if isinstance(v, (list, dict)): v = json.dumps(v, ensure_ascii=False)
                elif v is None: v = ""
                vals.append(v)
            writer.writerow(vals)
    return str(filepath)


def export_json(data: list[dict], filename: Optional[str] = None) -> str:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    if filename is None:
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    filepath = EXPORT_DIR / filename
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return str(filepath)
